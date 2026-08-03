"""
One-off exploratory analysis for the daily-target ask (Chris Roeder,
EVP Supply Chain & Ops, 2026-07-28) — NOT part of the production
McMaster pipeline, reads only from the existing daily archive
(reports/mcmaster/archive/), makes no database changes.

Question: for each branch, how does the realistically-achievable ceiling
of shippable lines (material available, order old enough to have had a
fair chance to be worked) change as the required minimum order age
("lead time") increases — and where does that ceiling cross below the
branch's target?

Reconstructs history from the archived reports rather than the database,
since mart_mcmaster__material_availability_daily only started logging
2026-07-28 (the underlying mcm_status has no history of its own and
can't be backfilled) — but every archived report's open_backlog_detail
tab already has real line-level order_dt + so_status going back to
whenever archiving started.

Outputs one workbook: reports/mcmaster/exploratory/target_lead_time_mockup.xlsx
  - raw_data:            every eligible line, one row each, across all archives
  - age_buckets_by_day:  org x archive_date, count of eligible lines per age bucket
  - lead_time_scenario:  the chart — small multiples, one panel per target org
"""
import glob
import os
import re

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ARCHIVE_DIR = os.path.join("reports", "mcmaster", "archive")
OUTPUT_DIR = os.path.join("reports", "mcmaster", "exploratory")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "target_lead_time_mockup.xlsx")
CHART_PATH = os.path.join(OUTPUT_DIR, "lead_time_scenario.png")

TARGETS = {"CLE": 70, "JVL": 60, "ATL": 30, "LOS": 25, "WIE": 25}  # DAL pending
SPHERE_ORGS = ["ATL", "CLE", "DAL", "JVL", "LOS", "WIE"]

AGE_BUCKET_EDGES = [-1, 0, 1, 2, 5, 10, 20, np.inf]
AGE_BUCKET_LABELS = ["0 days", "1 day", "2 days", "3-5 days", "6-10 days", "11-20 days", "21+ days"]

LEAD_TIMES = [0, 1, 2, 3, 5, 7, 10, 14]

# Same validated categorical/status colors already in use elsewhere in this
# report (mcmaster_output.py) — reused here for consistency across the
# report family, not re-picked.
COLOR_CEILING = "#52514e"    # neutral secondary ink — a measure, not a judgment
COLOR_TARGET = "#0b0b0b"     # near-black dashed reference line — a threshold, not a series
COLOR_AXIS = "#c3c2b7"
COLOR_GRID = "#e1e0d9"
COLOR_TEXT_SECONDARY = "#52514e"


def extract_all_archives():
    """One row per eligible (not 'No Material') McMaster line, per archive
    snapshot. age_days = archive_date - order_dt, always >= 0 since a line
    can't be ordered after the snapshot that shows it still open."""
    rows = []
    for path in sorted(glob.glob(os.path.join(ARCHIVE_DIR, "mcmaster_report_*.xlsx"))):
        m = re.search(r"(\d{4})_(\d{2})_(\d{2})", os.path.basename(path))
        archive_date = pd.Timestamp(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["open_backlog_detail"]
        headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        org_idx = headers.index("org")
        order_dt_idx = headers.index("order_dt")
        status_idx = headers.index("so_status")

        for row in ws.iter_rows(min_row=3, values_only=True):
            org = row[org_idx]
            order_dt = row[order_dt_idx]
            status = row[status_idx]
            if org not in SPHERE_ORGS or status == "No Material" or order_dt is None:
                continue
            age_days = (archive_date - order_dt.date()).days
            rows.append((archive_date, org, order_dt.date(), status, age_days))
        wb.close()

    df = pd.DataFrame(rows, columns=["archive_date", "org", "order_dt", "so_status", "age_days"])
    df["age_bucket"] = pd.cut(df["age_days"], bins=AGE_BUCKET_EDGES, labels=AGE_BUCKET_LABELS)
    return df


def build_age_buckets_by_day(df):
    pivot = df.pivot_table(
        index=["org", "archive_date"], columns="age_bucket", values="age_days",
        aggfunc="count", fill_value=0, observed=False,
    )
    return pivot[AGE_BUCKET_LABELS].reset_index()


def build_lead_time_scenario(df):
    """For each org/archive_date/lead_time L: count of eligible lines with
    age_days >= L (the material-driven ceiling if orders younger than L
    days don't count as fair-game). Averaged across archive dates per org."""
    records = []
    for org in TARGETS:
        org_df = df[df["org"] == org]
        for archive_date, day_df in org_df.groupby("archive_date"):
            for lt in LEAD_TIMES:
                records.append((org, archive_date, lt, (day_df["age_days"] >= lt).sum()))
    scenario = pd.DataFrame(records, columns=["org", "archive_date", "lead_time", "ceiling"])
    summary = (
        scenario.groupby(["org", "lead_time"])["ceiling"]
        .agg(avg="mean", min="min", max="max")
        .reset_index()
    )
    return scenario, summary


def render_chart(summary, n_archive_days):
    orgs = list(TARGETS.keys())
    fig, axes = plt.subplots(1, len(orgs), figsize=(18, 4.2), dpi=200, sharex=True)
    fig.patch.set_facecolor("#fcfcfb")

    for ax, org in zip(axes, orgs):
        d = summary[summary["org"] == org].sort_values("lead_time")
        ax.set_facecolor("#fcfcfb")
        ax.fill_between(d["lead_time"], d["min"], d["max"], color=COLOR_CEILING, alpha=0.15, linewidth=0)
        ax.plot(d["lead_time"], d["avg"], color=COLOR_CEILING, linewidth=2, marker="o", markersize=4,
                label="Avg. achievable ceiling")
        ax.axhline(TARGETS[org], color=COLOR_TARGET, linewidth=1.6, linestyle="--", label="Target")

        ax.set_title(f"{org}  (target {TARGETS[org]})", fontsize=11, color="#0b0b0b", loc="left")
        ax.set_xlabel("Min. order age required (days)", fontsize=8.5, color=COLOR_TEXT_SECONDARY)
        ax.grid(axis="y", color=COLOR_GRID, linewidth=0.8, zorder=0)
        ax.set_axisbelow(True)
        for spine in ("top", "right", "left"):
            ax.spines[spine].set_visible(False)
        ax.spines["bottom"].set_color(COLOR_AXIS)
        ax.tick_params(colors=COLOR_TEXT_SECONDARY, labelsize=8)
        ax.set_ylim(bottom=0)

    axes[0].set_ylabel("Lines (material available, old enough to count)", fontsize=8.5, color=COLOR_TEXT_SECONDARY)
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 1.06), ncol=2, frameon=False, fontsize=9)
    fig.suptitle(
        f"Achievable shipping ceiling vs. target, by required minimum order age  "
        f"(avg. + range across {n_archive_days} sampled days, 2026-07-07 to 2026-07-28)",
        fontsize=10.5, color="#0b0b0b", y=1.14,
    )
    fig.tight_layout()
    fig.savefig(CHART_PATH, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def write_workbook(df, buckets_df, scenario_df, summary_df):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "raw_data"
    ws1.append(["archive_date", "org", "order_dt", "so_status", "age_days", "age_bucket"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for row in df.itertuples(index=False):
        ws1.append([row.archive_date, row.org, row.order_dt, row.so_status, row.age_days, str(row.age_bucket)])

    ws2 = wb.create_sheet("age_buckets_by_day")
    ws2.append(list(buckets_df.columns))
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for row in buckets_df.itertuples(index=False):
        ws2.append(list(row))

    ws3 = wb.create_sheet("lead_time_scenario")
    ws3.append(["org", "target", "lead_time_days", "avg_ceiling", "min_ceiling", "max_ceiling"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for row in summary_df.itertuples(index=False):
        ws3.append([row.org, TARGETS[row.org], row.lead_time, row.avg, row.min, row.max])
    img = XLImage(CHART_PATH)
    ws3.add_image(img, f"A{len(summary_df) + 3}")

    for ws in (ws1, ws2, ws3):
        for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(col[0].value)) + 2)

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    print("Extracting from archives...")
    df = extract_all_archives()
    n_days = df["archive_date"].nunique()
    print(f"  {len(df)} eligible lines across {n_days} archive dates")

    buckets_df = build_age_buckets_by_day(df)
    scenario_df, summary_df = build_lead_time_scenario(df)

    print("Rendering chart...")
    render_chart(summary_df, n_days)

    print("Writing workbook...")
    write_workbook(df, buckets_df, scenario_df, summary_df)
    print(f"Done: {OUTPUT_PATH}")
