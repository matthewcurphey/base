"""
One-off preview export for mart_mcmaster__daily_target_performance — NOT
wired into the production mcmaster_output.py pipeline. The metric tab
itself may not go live in the template for a few days yet, so this just
dumps the mart's current columns to a plain workbook to copy-paste into
the hand-built template tab while its design is still being tinkered with.

Data will show nulls for variance_to_target/pct_to_target on 07-29 and
earlier — the material-availability log was schema-migrated (lt_bucket
added) and lost its pre-07-30 history in the process, so there's no
material snapshot to score those days against yet. Expected, not a bug.

Run: python reports/mcmaster/exploratory/target_performance_preview.py
"""
import os

import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from etl.utils.connect_postgres import get_postgres_connection

OUTPUT_PATH = os.path.join("reports", "mcmaster", "exploratory", "target_performance_preview.xlsx")


def main():
    engine = get_postgres_connection()
    df = pd.read_sql(
        "SELECT * FROM analytics_marts.mart_mcmaster__daily_target_performance "
        "ORDER BY dt DESC, inv_org_code ASC",
        engine,
    )
    engine.dispose()

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "daily_target_performance"

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in df.itertuples(index=False):
        ws.append(list(row))

    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    wb.save(OUTPUT_PATH)
    print(f"{len(df)} rows written: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
