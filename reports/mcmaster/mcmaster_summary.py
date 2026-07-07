import os
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

# Fixed layout of the live template's summary sheet (hand-formatted;
# python only ever overwrites values here, never style). Unlike every
# other tab, this structure never grows or shrinks — always these 6 orgs,
# always 7 days, always these 4 statuses — so there's no leftover-row
# clearing logic needed, unlike the mart-driven tabs.
HISTORICAL_VIEW_CELL = "A1"
LIVE_VIEW_CELL = "M1"
TREND_ANCHOR = "A3"
TREND_RANGE = (1, 10, 3, 29)          # min_col, max_col, min_row, max_row
STATUS_CHART_ANCHOR = "M3"
STATUS_CHART_RANGE = (13, 22, 3, 19)
SEVEN_DAY_DOW_ROW = 31
SEVEN_DAY_HEADER_ROW = 32
SEVEN_DAY_FIRST_DATA_ROW = 33
SEVEN_DAY_COL = 1                     # A
STATUS_PIVOT_HEADER_ROW = 21
STATUS_PIVOT_COL = 13                 # M

CHART_PATH = os.path.join("reports", "mcmaster", "backlog_trend.png")

# McMaster's 6-org sphere — consistent with every other mcmaster model/mart.
# CHA/PHI/STO have had zero McMaster activity for 7+ months (not dormant this
# week specifically, just not part of this business) and would only add rows
# of permanent zeros to an exec-facing table.
SPHERE_ORGS = ["ATL", "CLE", "DAL", "JVL", "LOS", "WIE"]

# Palette slots from the dataviz skill's reference palette (validated set),
# not arbitrary picks. Red/green here also carry a positional redundancy —
# new sits above the zero line, shipped sits below — so identity doesn't
# depend on color discrimination alone. Backlog is neutral (secondary ink),
# not a categorical hue — it's the scale/context the red/green signal reads
# against, not a series competing for identity.
COLOR_BACKLOG = "#52514e"   # secondary ink (neutral grey)
COLOR_NEW = "#e34948"       # categorical slot 6 (red)
COLOR_SHIPPED = "#008300"   # categorical slot 4 (green)
COLOR_AXIS = "#c3c2b7"
COLOR_GRID = "#e1e0d9"
COLOR_TEXT_SECONDARY = "#52514e"


def _col_width_px(ws, col_idx):
    """Excel column width (character units) to pixels — standard Calibri 11
    approximation (width*7 + 5), which is what openpyxl assumes at 96 DPI."""
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    width = dim.width if dim and dim.width else (ws.sheet_format.defaultColWidth or 8.43)
    return width * 7 + 5


def _row_height_px(ws, row_idx):
    """Excel row height (points) to pixels at 96 DPI (pt * 96/72)."""
    dim = ws.row_dimensions.get(row_idx)
    height = dim.height if dim and dim.height else (ws.sheet_format.defaultRowHeight or 15)
    return height * 4 / 3


def range_size_px(ws, min_col, max_col, min_row, max_row):
    """
    Pixel size of a cell range as openpyxl/Excel will actually render it.
    openpyxl does NOT read a PNG's DPI metadata when placing an image — it
    takes the raw pixel dimensions and assumes 96 DPI. So "make the chart
    14x5 inches" via matplotlib's figsize/dpi does not control the size it
    displays at in Excel; only the image's actual pixel count does. This is
    why the first cut rendered far larger than intended (~22x7.8in instead
    of 14x5). Use this to compute the real pixel target for a cell range,
    then force the embedded image to exactly that pixel size.
    """
    width = sum(_col_width_px(ws, c) for c in range(min_col, max_col + 1))
    height = sum(_row_height_px(ws, r) for r in range(min_row, max_row + 1))
    return round(width), round(height)


def build_trend_chart(backlog_daily_df, days=90, output_path=CHART_PATH, figsize=(9.6, 8.1), dpi=200):
    """
    90-day (default) company-wide trend: daily backlog level as bars, and
    daily new/shipped orders as a thin area straddling zero on the SAME
    axis — new above zero, shipped below (plotted as negative so it renders
    underneath the line). Raw daily values, no smoothing: "this is what
    actually happened," weekend zeros included as-is.

    figsize/dpi control the saved file's resolution (crispness) only — the
    size it displays at in Excel is set separately in update_live_summary_sheet
    via explicit pixel dimensions (see range_size_px).
    """
    df = backlog_daily_df[backlog_daily_df["inv_org_code"].isin(SPHERE_ORGS)]

    daily_total = (
        df.groupby("dt", as_index=False)[["open_orders", "new_orders", "shipped_orders"]]
        .sum()
        .sort_values("dt")
    )
    daily_total = daily_total.tail(days)

    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    fig.patch.set_facecolor("#fcfcfb")
    ax.set_facecolor("#fcfcfb")

    ax.bar(
        daily_total["dt"], daily_total["open_orders"],
        color=COLOR_BACKLOG, width=0.9, label="Backlog (open orders)", zorder=2,
    )
    ax.fill_between(
        daily_total["dt"], 0, daily_total["new_orders"],
        color=COLOR_NEW, alpha=0.85, linewidth=0, label="New orders", zorder=3,
    )
    ax.fill_between(
        daily_total["dt"], 0, -daily_total["shipped_orders"],
        color=COLOR_SHIPPED, alpha=0.85, linewidth=0, label="Shipped", zorder=3,
    )

    ax.axhline(0, color=COLOR_AXIS, linewidth=1)
    ax.grid(axis="y", color=COLOR_GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)

    for spine in ("top", "right", "left"):
        ax.spines[spine].set_visible(False)
    ax.spines["bottom"].set_color(COLOR_AXIS)

    ax.tick_params(colors=COLOR_TEXT_SECONDARY, labelsize=9)
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.MO, interval=2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%d-%b"))
    fig.autofmt_xdate(rotation=0, ha="center")

    fig.suptitle(
        f"McMaster Open Backlog — {days}-Day Trend",
        fontsize=13, color="#0b0b0b", x=0.01, ha="left", y=0.98,
    )
    handles, labels = ax.get_legend_handles_labels()
    order = [labels.index(name) for name in ("Backlog (open orders)", "New orders", "Shipped")]
    ax.legend(
        [handles[i] for i in order], [labels[i] for i in order],
        loc="upper center", bbox_to_anchor=(0.5, -0.12), ncol=3,
        frameon=False, fontsize=9, labelcolor=COLOR_TEXT_SECONDARY,
    )

    fig.tight_layout(rect=(0, 0.04, 1, 0.93))
    fig.savefig(output_path, facecolor=fig.get_facecolor())
    plt.close(fig)

    return output_path


def get_7day_window(backlog_daily_df):
    """
    Last 7 calendar days ending at the mart's most recent date (yesterday,
    same date basis the pipeline already uses) — a fixed window, weekends
    included even when empty. Returns dates oldest-to-newest (day-7..day-1).
    """
    end_date = backlog_daily_df["dt"].max()
    return [end_date - timedelta(days=i) for i in range(6, -1, -1)]


def build_7day_activity_table(backlog_daily_df):
    """
    One 3-row block per sphere org (backlog/new/shipped) plus a Total block,
    each showing the 7 raw daily values (fixed calendar window, weekends
    included as-is — no smoothing, "what actually happened") and the
    current 5-day MA pulled straight from the mart's own _5d_avg columns,
    not recalculated. Total's MA is the sum of the 6 orgs' MAs — the mart
    has no company-wide MA row of its own to pull from.

    Returns (dates, rows) where rows is a list of
    (org_label, metric_label, [7 daily values], ma_value).
    """
    dates = get_7day_window(backlog_daily_df)
    end_date = dates[-1]

    window = backlog_daily_df[
        backlog_daily_df["inv_org_code"].isin(SPHERE_ORGS) & backlog_daily_df["dt"].isin(dates)
    ]

    metrics = [
        ("backlog", "open_orders", "backlog_5d_avg"),
        ("new", "new_orders", "new_orders_5d_avg"),
        ("shipped", "shipped_orders", "shipped_orders_5d_avg"),
    ]

    rows = []
    for org in SPHERE_ORGS:
        org_df = window[window["inv_org_code"] == org].set_index("dt")
        for metric_label, value_col, ma_col in metrics:
            daily_values = [org_df.loc[d, value_col] if d in org_df.index else 0 for d in dates]
            ma_value = org_df.loc[end_date, ma_col] if end_date in org_df.index else None
            rows.append((org, metric_label, daily_values, ma_value))

    for metric_label, value_col, ma_col in metrics:
        daily_values = [window.loc[window["dt"] == d, value_col].sum() for d in dates]
        ma_value = window.loc[window["dt"] == end_date, ma_col].sum()
        rows.append(("Total", metric_label, daily_values, ma_value))

    return dates, rows


# True fulfillment lifecycle order (from the mcm_status SQL logic), not
# alphabetical: blocked -> ready -> in production -> done.
STATUS_ORDER = ["No Material", "Material Available", "Job Started", "Job Complete"]

# Matches the conditional-formatting colors already used for this exact
# status column on open_backlog_detail (AV) — same status, same color,
# everywhere in the report. "Material Available" is a theme color there
# (theme index 3 = Dark 2, #0E2841, tint 0.75) resolved to its actual RGB.
STATUS_COLORS = {
    "No Material": "#FFC7CE",
    "Material Available": "#A6CAEC",
    "Job Started": "#FFEB9C",
    "Job Complete": "#C6EFCE",
}


def build_backlog_status_pivot(backlog_status_df):
    """Org rows x status columns, line_count values, totals row/column —
    columns ordered by pipeline stage, not alphabetically."""
    pivot = backlog_status_df.pivot_table(
        index="inv_org_code", columns="mcm_status", values="line_count",
        aggfunc="sum", fill_value=0, margins=True, margins_name="Total",
    )
    ordered_cols = [c for c in STATUS_ORDER if c in pivot.columns] + ["Total"]
    return pivot[ordered_cols]


def build_status_chart(backlog_status_df, output_path=None, figsize=(9.6, 4.3), dpi=200):
    """
    Company-wide (no org breakdown) horizontal bar, one bar per pipeline
    stage in true lifecycle order, single-hue light->dark ramp so the
    sequence reads through position AND color. Direct value labels — only
    4 categories, no legend needed.
    """
    if output_path is None:
        output_path = os.path.join("reports", "mcmaster", "status_chart.png")

    totals = backlog_status_df.groupby("mcm_status")["line_count"].sum()
    totals = totals.reindex([s for s in STATUS_ORDER if s in totals.index])

    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    fig.patch.set_facecolor("#fcfcfb")
    ax.set_facecolor("#fcfcfb")

    y_pos = range(len(totals))
    bars = ax.barh(
        y_pos, totals.values,
        color=[STATUS_COLORS[s] for s in totals.index], height=0.6, zorder=2,
    )
    ax.set_yticks(list(y_pos))
    ax.set_yticklabels(totals.index, fontsize=10, color="#0b0b0b")
    ax.invert_yaxis()  # earliest stage on top, reads top-to-bottom as progression

    max_val = totals.values.max()
    for bar, value in zip(bars, totals.values):
        ax.text(
            bar.get_width() + max_val * 0.02, bar.get_y() + bar.get_height() / 2,
            f"{int(value):,}", va="center", ha="left", fontsize=10, color="#0b0b0b",
        )

    ax.set_xlim(0, max_val * 1.15)
    ax.set_xticks([])
    for spine in ("top", "right", "bottom", "left"):
        ax.spines[spine].set_visible(False)

    fig.suptitle(
        "McMaster Open Backlog by Status — Company-Wide", fontsize=13, color="#0b0b0b",
        x=0.01, ha="left", y=0.98,
    )

    fig.tight_layout(rect=(0, 0.02, 1, 0.90))
    fig.savefig(output_path, facecolor=fig.get_facecolor())
    plt.close(fig)

    return output_path


def _aspect_figsize(w_px, h_px, base=9.0):
    """
    Figsize preserving the given pixel aspect ratio, scaled to a fixed
    physical size chosen for good font-to-chart proportions — matplotlib
    font sizes are in points (tied to physical inches), not pixels, so a
    tiny figsize makes text look oversized regardless of dpi. The actual
    target pixel count is applied separately via explicit width/height on
    the embedded image (_place_image), fully decoupled from this.
    """
    ratio = w_px / h_px
    if ratio >= 1:
        return (base, base / ratio)
    return (base * ratio, base)


def _place_image(ws, path, anchor_cell, min_col, max_col, min_row, max_row):
    """
    Embed an image sized to exactly fill the given cell range. openpyxl
    ignores a PNG's DPI metadata when placing it — it uses raw pixel count
    at an assumed 96 DPI — so the source file's resolution only controls
    crispness; display size has to be set explicitly here.
    """
    target_w, target_h = range_size_px(ws, min_col, max_col, min_row, max_row)
    img = XLImage(path)
    img.width = target_w
    img.height = target_h
    img.anchor = anchor_cell
    ws.add_image(img)
    return target_w, target_h


def _replace_image(ws, path, anchor_cell, min_col, max_col, min_row, max_row):
    """
    Same sizing as _place_image, but first drops any existing image already
    anchored at this exact cell — otherwise every run adds another image on
    top of the last one instead of replacing it.
    """
    row, col = coordinate_to_tuple(anchor_cell)
    target_row0, target_col0 = row - 1, col - 1
    ws._images = [
        img for img in ws._images
        if not (hasattr(img.anchor, "_from")
                and img.anchor._from.row == target_row0
                and img.anchor._from.col == target_col0)
    ]
    return _place_image(ws, path, anchor_cell, min_col, max_col, min_row, max_row)


def update_live_summary_sheet(wb, backlog_daily_df, backlog_status_df):
    """
    Refresh the summary sheet in the live, hand-formatted template —
    values only, at fixed known positions. Formatting is never touched
    here; the sheet was styled by hand and this must not disturb it.

    Left side (A) is the historical view: complete-days-only trend chart
    and 7-day activity table, both from backlog_daily. Right side (M) is
    the live view: current-snapshot status chart and pivot, both from
    backlog_status, stamped with the run timestamp so it's clear this
    side reflects the moment the report was generated, not settled history.
    """
    ws = wb["summary"]

    last_complete_date = backlog_daily_df["dt"].max()
    ws[HISTORICAL_VIEW_CELL] = f"Historical View — Complete Days Through {last_complete_date:%d-%b-%Y}"
    ws[LIVE_VIEW_CELL] = f"Live View — As Of {datetime.now():%d-%b-%Y %H:%M}"

    trend_w, trend_h = range_size_px(ws, TREND_RANGE[0], TREND_RANGE[1], TREND_RANGE[2], TREND_RANGE[3])
    trend_path = build_trend_chart(backlog_daily_df, figsize=_aspect_figsize(trend_w, trend_h), dpi=200)
    _replace_image(ws, trend_path, TREND_ANCHOR, *TREND_RANGE)

    status_w, status_h = range_size_px(
        ws, STATUS_CHART_RANGE[0], STATUS_CHART_RANGE[1], STATUS_CHART_RANGE[2], STATUS_CHART_RANGE[3]
    )
    status_path = build_status_chart(backlog_status_df, figsize=_aspect_figsize(status_w, status_h), dpi=200)
    _replace_image(ws, status_path, STATUS_CHART_ANCHOR, *STATUS_CHART_RANGE)

    dates, rows = build_7day_activity_table(backlog_daily_df)
    for i, d in enumerate(dates):
        col = SEVEN_DAY_COL + 2 + i
        ws.cell(SEVEN_DAY_DOW_ROW, col, d.strftime("%a"))
        ws.cell(SEVEN_DAY_HEADER_ROW, col, d)
    ma_col = SEVEN_DAY_COL + 2 + len(dates)
    for j, (org, metric_label, daily_values, ma_value) in enumerate(rows):
        row = SEVEN_DAY_FIRST_DATA_ROW + j
        if metric_label == "backlog":
            ws.cell(row, SEVEN_DAY_COL, org)
        ws.cell(row, SEVEN_DAY_COL + 1, metric_label)
        for i, v in enumerate(daily_values):
            ws.cell(row, SEVEN_DAY_COL + 2 + i, None if pd.isna(v) else v)
        ws.cell(row, ma_col, None if pd.isna(ma_value) else ma_value)

    pivot = build_backlog_status_pivot(backlog_status_df)
    for j, status in enumerate(pivot.columns):
        ws.cell(STATUS_PIVOT_HEADER_ROW, STATUS_PIVOT_COL + 1 + j, status)
    for i, org in enumerate(pivot.index):
        row = STATUS_PIVOT_HEADER_ROW + 1 + i
        ws.cell(row, STATUS_PIVOT_COL, org)
        for j, status in enumerate(pivot.columns):
            ws.cell(row, STATUS_PIVOT_COL + 1 + j, int(pivot.loc[org, status]))

    return ws
